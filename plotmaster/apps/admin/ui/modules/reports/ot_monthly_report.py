import threading
import customtkinter as ctk
from tkinter import messagebox, filedialog
from datetime import datetime, date, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except Exception:
    _OPENPYXL_AVAILABLE = False

try:
    from tkcalendar import DateEntry  # Mejora UX si está disponible
except Exception:
    DateEntry = None

from plotmaster.core.services.supabase_service import get_work_orders_between


class BaseReporteExcelOT(ctk.CTkFrame):
    """Componente reutilizable para exportar OTs a Excel con filtros estándar."""

    def __init__(
        self,
        parent,
        *,
        title: str,
        subtitle: str,
        status_filter: str = None,
        forma_pago_filter: str = None,
        filename_prefix: str = "reporte_ot",
        filename_suffix: str = "",
        sheet_title: str = "OT",
        empty_message: str = "Sin datos en el rango indicado.",
        **kwargs,
    ):
        super().__init__(parent, fg_color="transparent", **kwargs)
        self.title_text = title
        self.subtitle_text = subtitle
        self.status_filter = status_filter
        self.forma_pago_filter = forma_pago_filter
        self.filename_prefix = filename_prefix
        self.filename_suffix = filename_suffix
        self.sheet_title = sheet_title
        self.empty_message = empty_message
        self._build_ui()

    # UI -----------------------------------------------------------------
    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)

        header = ctk.CTkLabel(
            self,
            text=self.title_text,
            font=ctk.CTkFont(size=22, weight="bold"),
        )
        header.grid(row=0, column=0, sticky="w", padx=10, pady=(10, 6))

        subtitle = ctk.CTkLabel(
            self,
            text=self.subtitle_text,
            font=ctk.CTkFont(size=13),
            text_color="#4b5563",
            wraplength=720,
            justify="left",
        )
        subtitle.grid(row=1, column=0, sticky="w", padx=10, pady=(0, 12))

        card = ctk.CTkFrame(self, fg_color="#FFFFFF", corner_radius=12, border_width=1, border_color="#E5E7EB")
        card.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")
        card.grid_columnconfigure((0, 1, 2), weight=1)

        self.fecha_desde_widget = self._create_date_input(card, "Fecha desde", 0)
        self.fecha_hasta_widget = self._create_date_input(card, "Fecha hasta", 1)

        rango_frame = ctk.CTkFrame(card, fg_color="transparent")
        rango_frame.grid(row=0, column=2, padx=10, pady=10, sticky="nsew")
        ctk.CTkLabel(rango_frame, text="Atajos de rango", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w")
        self.rango_var = ctk.StringVar(value="personalizado")
        self.rango_picker = ctk.CTkSegmentedButton(
            rango_frame,
            values=["Mes actual", "Últimos 30 días", "Personalizado"],
            variable=self.rango_var,
            command=self._on_rango_change,
        )
        self.rango_picker.pack(fill="x", pady=(8, 2))
        ctk.CTkLabel(rango_frame, text="Aplicamos las fechas automáticamente.", font=ctk.CTkFont(size=11), text_color="#6b7280").pack(anchor="w")

        orden_frame = ctk.CTkFrame(card, fg_color="transparent")
        orden_frame.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="nsew")
        ctk.CTkLabel(orden_frame, text="Orden por fecha", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w")
        self.orden_var = ctk.StringVar(value="desc")
        self.orden_picker = ctk.CTkSegmentedButton(orden_frame, values=["asc", "desc"], variable=self.orden_var)
        self.orden_picker.pack(fill="x", pady=(8, 0))
        ctk.CTkLabel(orden_frame, text="Desc: más recientes primero", font=ctk.CTkFont(size=11), text_color="#6b7280").pack(anchor="w", pady=(4, 0))

        self.btn_descargar = ctk.CTkButton(
            card,
            text="📊 Exportar Excel",
            fg_color="#2563EB",
            hover_color="#1D4ED8",
            command=self._on_descargar,
            height=42,
            corner_radius=8,
        )
        self.btn_descargar.grid(row=1, column=1, columnspan=2, padx=10, pady=(0, 12), sticky="ew")

        self.status_label = ctk.CTkLabel(card, text="", text_color="#374151")
        self.status_label.grid(row=2, column=0, columnspan=3, padx=10, pady=(0, 10), sticky="w")

    def _create_date_input(self, parent, label, col):
        wrapper = ctk.CTkFrame(parent, fg_color="transparent")
        wrapper.grid(row=0, column=col, padx=10, pady=10, sticky="nsew")
        ctk.CTkLabel(wrapper, text=label, font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w")
        input_card = ctk.CTkFrame(wrapper, fg_color="#f7f9fb", corner_radius=10, border_width=1, border_color="#e5e7eb")
        input_card.pack(fill="x", pady=(6, 0))

        row = ctk.CTkFrame(input_card, fg_color="transparent")
        row.pack(fill="x", padx=10, pady=(8, 4))
        ctk.CTkLabel(row, text="📅", width=28, anchor="center", font=ctk.CTkFont(size=14)).pack(side="left", padx=(0, 8))

        if DateEntry:
            picker = DateEntry(row, width=14, date_pattern='yyyy-mm-dd')
            picker.set_date(date.today())
        else:
            picker = ctk.CTkEntry(row, placeholder_text="yyyy-mm-dd")
            picker.insert(0, date.today().isoformat())
        picker.pack(side="left", fill="x", expand=True)

        ctk.CTkLabel(input_card, text="Formato: yyyy-mm-dd", font=ctk.CTkFont(size=10), text_color="#6b7280").pack(anchor="w", padx=10, pady=(0, 8))
        return picker

    # Lógica --------------------------------------------------------------
    def _on_rango_change(self, selected):
        hoy = date.today()
        if selected == "Mes actual":
            inicio = hoy.replace(day=1)
        elif selected == "Últimos 30 días":
            inicio = hoy - timedelta(days=30)
        else:
            return
        self._set_date_widget(self.fecha_desde_widget, inicio)
        self._set_date_widget(self.fecha_hasta_widget, hoy)

    def _set_date_widget(self, widget, value_date: date):
        if DateEntry and isinstance(widget, DateEntry):
            widget.set_date(value_date)
        else:
            widget.delete(0, 'end')
            widget.insert(0, value_date.isoformat())

    def _parse_date(self, widget):
        if DateEntry and isinstance(widget, DateEntry):
            return widget.get_date()
        raw = widget.get().strip()
        return datetime.strptime(raw, "%Y-%m-%d").date()

    def _on_descargar(self):
        try:
            f_ini = self._parse_date(self.fecha_desde_widget)
            f_fin = self._parse_date(self.fecha_hasta_widget)
        except Exception:
            messagebox.showerror("Fechas inválidas", "Usa el formato yyyy-mm-dd para ambas fechas.")
            return

        if f_ini > f_fin:
            messagebox.showerror("Rango inválido", "La fecha inicial no puede ser mayor que la fecha final.")
            return

        if not _OPENPYXL_AVAILABLE:
            messagebox.showerror("Dependencia faltante", "Instala 'openpyxl' para exportar a Excel.")
            return

        sort_desc = self.orden_var.get() == "desc"
        self.status_label.configure(text="Consultando datos…", text_color="#374151")
        self.btn_descargar.configure(state="disabled", text="Consultando…")
        threading.Thread(target=self._fetch_data_async, args=(f_ini, f_fin, sort_desc), daemon=True).start()

    def _fetch_data_async(self, f_ini, f_fin, sort_desc):
        ok, data = get_work_orders_between(
            f_ini,
            f_fin,
            sort_desc=sort_desc,
            status=self.status_filter,
            forma_pago=self.forma_pago_filter,
        )
        self.after(0, lambda: self._after_data_fetch(ok, data, f_ini, f_fin))

    def _after_data_fetch(self, ok, data, f_ini, f_fin):
        self.btn_descargar.configure(state="normal", text="📊 Exportar Excel")
        if not ok:
            messagebox.showerror("Error", data)
            self.status_label.configure(text="", text_color="#b91c1c")
            return

        filtered = [r for r in (data or []) if self._matches_filters(r)]
        if not filtered:
            messagebox.showinfo("Sin datos", self.empty_message)
            self.status_label.configure(text="")
            return

        suffix = f"_{self.filename_suffix}" if self.filename_suffix else ""
        default_name = f"{self.filename_prefix}_{f_ini.isoformat()}_{f_fin.isoformat()}{suffix}.xlsx"
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
        )
        if not path:
            self.status_label.configure(text="Exportación cancelada.", text_color="#6b7280")
            return

        try:
            self._write_excel(path, filtered)
            self.status_label.configure(text=f"Archivo guardado: {path}", text_color="#059669")
            messagebox.showinfo("Éxito", "Reporte Excel generado y listo para usar.")
        except Exception as e:
            messagebox.showerror("Error al escribir archivo", str(e))
            self.status_label.configure(text="No se pudo guardar el archivo.", text_color="#b91c1c")

    # Excel ---------------------------------------------------------------
    def _write_excel(self, path, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_title

        headers = [
            ("OT", 'ot_nro'),
            ("Fecha creación", 'fecha_creacion'),
            ("Cliente", 'cliente'),
            ("CI/RUC Cliente", 'cliente_ci_ruc'),
            ("Vendedor", 'vendedor'),
            ("CI/RUC Vendedor", 'vendedor_ci_ruc'),
            ("Descripción", 'descripcion'),
            ("Valor total", 'valor_total'),
            ("Abonado", 'abonado_total'),
            ("Forma de pago", 'forma_pago'),
            ("Envío", 'solicita_envio'),
            ("Fecha entrega", 'fecha_entrega'),
        ]

        ws.append([h[0] for h in headers])

        for r in rows:
            ws.append([
                r.get('ot_nro'),
                r.get('fecha_creacion'),
                r.get('cliente'),
                r.get('cliente_ci_ruc'),
                r.get('vendedor'),
                r.get('vendedor_ci_ruc'),
                r.get('descripcion'),
                float(r.get('valor_total') or 0),
                float(r.get('abonado_total') or 0),
                r.get('forma_pago'),
                'Con Envío' if r.get('solicita_envio') else 'Sin Envío',
                r.get('fecha_entrega') or '',
            ])

        header_fill = PatternFill("solid", fgColor="1D4ED8")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style="thin", color="E5E7EB"),
                             right=Side(style="thin", color="E5E7EB"),
                             top=Side(style="thin", color="E5E7EB"),
                             bottom=Side(style="thin", color="E5E7EB"))

        col_widths = {
            "A": 10,
            "B": 14,
            "C": 26,
            "D": 18,
            "E": 22,
            "F": 18,
            "G": 38,
            "H": 14,
            "I": 12,
            "J": 14,
            "K": 12,
            "L": 14,
        }

        for col_idx, (title, _) in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            ws.column_dimensions[col_letter].width = col_widths.get(col_letter, 16)

        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=2).number_format = "yyyy-mm-dd"
            ws.cell(row=row_idx, column=12).number_format = "yyyy-mm-dd"
            ws.cell(row=row_idx, column=8).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=9).number_format = "#,##0.00"
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_idx in (8, 9):
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in (1, 2, 12):
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")
                cell.border = thin_border

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        ws.freeze_panes = "A2"

        wb.save(path)

    # Utilidades ---------------------------------------------------------
    def _normalize_text(self, value):
        text = str(value or '').strip().lower()
        for orig, repl in (
            ('á', 'a'), ('é', 'e'), ('í', 'i'), ('ó', 'o'), ('ú', 'u'), ('ü', 'u'),
        ):
            text = text.replace(orig, repl)
        return text

    def _matches_filters(self, row: dict) -> bool:
        if self.status_filter:
            if self._normalize_text(row.get('status')) != self._normalize_text(self.status_filter):
                return False
        if self.forma_pago_filter:
            if self._normalize_text(row.get('forma_pago')) != self._normalize_text(self.forma_pago_filter):
                return False
        return True


class ReporteMensualOT(BaseReporteExcelOT):
    """Reporte mensual (OT finalizadas)."""

    def __init__(self, parent, **kwargs):
        super().__init__(
            parent,
            title="Reporte mensual",
            subtitle="Descarga un Excel listo para entregar.",
            status_filter="Finalizado",
            filename_prefix="reporte_ot",
            filename_suffix="FINALIZADAS",
            sheet_title="OT Finalizadas",
            empty_message="No hay órdenes finalizadas en el rango indicado.",
            **kwargs,
        )


class ReporteCreditoOT(BaseReporteExcelOT):
    """Reporte de crédito: entregadas y con forma de pago crédito."""

    def __init__(self, parent, **kwargs):
        super().__init__(
            parent,
            title="Reporte de Crédito",
            subtitle="Exporta solo OTs ENTREGADAS con forma de pago CRÉDITO.",
            status_filter="Entregado",
            forma_pago_filter="Crédito",
            filename_prefix="reporte_ot_credito",
            filename_suffix="CREDITO",
            sheet_title="OT Crédito",
            empty_message="No hay órdenes ENTREGADAS con pago CRÉDITO en el rango indicado.",
            **kwargs,
        )
